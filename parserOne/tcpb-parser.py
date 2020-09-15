import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.122 Safari/537.36',
           'accept': '*/*'}

def get_html(url, params=None):
    res = requests.get(url, headers=HEADERS, params=params)
    return res

def get_content(html):
    soup = BeautifulSoup(html, 'html.parser')
    content = soup.find_all('span', class_='uc-price')[0].text
    # for img in content('img'):
    #     img.decompose()
    return content


def parse(URL):
    html = get_html(URL)
    if html.status_code == 200:
        return str(get_content(html.text))
    else:
        return 'Error'


wb = load_workbook('../data/tcpb.xlsx')
ws = wb.active
urls_raw = ws['H']
content_raw = ws['I']

for i in range(1, 100):
    url = urls_raw[i].value
    content_raw[i].value = parse(url)

wb.save('output/tcpb.xlsx')


